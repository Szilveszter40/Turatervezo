import json
import os
from openpyxl import load_workbook
from PyQt6.QtWidgets import (QPushButton, QDialog, QVBoxLayout, QHBoxLayout, 
                             QLabel, QMessageBox, QLineEdit, QComboBox, 
                             QTableWidget, QTableWidgetItem, QHeaderView, 
                             QTreeWidgetItem, QListWidget, QTabWidget, QWidget, QFileDialog, QFrame)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QFont

class ModulInit:
    def __init__(self, main_app):
        self.main = main_app
        self.mappa = os.path.dirname(os.path.abspath(__file__))
        
        # Fájl elérési utak
        self.adat_fajl = os.path.join(self.mappa, "tervezo_partnerek.json")
        self.tura_adatok_fajl = os.path.join(self.mappa, "tervezo_turak_mentese.json")
        self.turanevek_txt = os.path.join(os.getcwd(), "Turanevek.txt")
        
        self.fejlecek = ["Név", "KI", "BE", "Db", "Kg", "Megjegyzés", "Város", "Irsz"]
        
        self.betolt_adatok()
        
        # Késleltetett indítás a tiszta felületért
        QTimer.singleShot(1000, self.init_gomb)

    def takaritas(self):
        """Eltávolítja a régi Tervező gombokat a duplikáció elkerülésére"""
        if hasattr(self.main, 'tervezo_btn_obj'):
            try:
                old_btn = getattr(self.main, 'tervezo_btn_obj')
                old_btn.setParent(None)
                old_btn.deleteLater()
            except: pass

    def betolt_adatok(self):
        try:
            if os.path.exists(self.adat_fajl):
                with open(self.adat_fajl, "r", encoding="utf-8") as f: self.partnerek = json.load(f)
            else: self.partnerek = []
        except: self.partnerek = []
        try:
            if os.path.exists(self.tura_adatok_fajl):
                with open(self.tura_adatok_fajl, "r", encoding="utf-8") as f: self.tura_kapcsolatok = json.load(f)
            else: self.tura_kapcsolatok = {}
        except: self.tura_kapcsolatok = {}

    def mentes_adatok(self):
        with open(self.adat_fajl, "w", encoding="utf-8") as f: json.dump(self.partnerek, f, indent=4)
        with open(self.tura_adatok_fajl, "w", encoding="utf-8") as f: json.dump(self.tura_kapcsolatok, f, indent=4)

    def init_gomb(self):
        self.takaritas()
        self.main.tervezo_btn_obj = QPushButton("🏗️ TERVEZŐ")
        self.main.tervezo_btn_obj.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71; 
                color: white; 
                font-weight: bold; 
                padding: 6px; 
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        self.main.tervezo_btn_obj.clicked.connect(self.megjelenites)

        if hasattr(self.main, 'left_tree'):
            layout = self.main.left_tree.parent().layout()
            layout.insertWidget(2, self.main.tervezo_btn_obj)

    def turak_behelyezese_a_foprogramba(self):
        if not hasattr(self.main, 'add_tura_item') or not hasattr(self.main, 'right_tree'): return
        
        self.main.right_tree.blockSignals(True)
        count = 0
        try:
            for t_nev, p_nevek in self.tura_kapcsolatok.items():
                if not p_nevek: continue
                
                ti = self.main.add_tura_item(f"🚛 {t_nev}")
                ti.setData(0, Qt.ItemDataRole.UserRole, "TURA")
                
                for p_nev in p_nevek:
                    p = next((x for x in self.partnerek if x['nev'] == p_nev), None)
                    if not p: continue
                    
                    ex = QTreeWidgetItem(ti, [p['nev'], p.get('ki',''), p.get('be',''), str(p.get('db','0')), str(p.get('kg','0')), p.get('megj',''), p.get('varos',''), p.get('irsz','')])
                    ex.setData(0, Qt.ItemDataRole.UserRole, "PARTNER")
                    QTreeWidgetItem(ex, ["", p.get('ki',''), p.get('be',''), str(p.get('db','0')), str(p.get('kg','0')), "", "", ""])
                
                ti.setExpanded(True)
                count += 1
            
            if hasattr(self.main, 'recalculate'):
                self.main.recalculate()
                
            QMessageBox.information(self.dialog, "Siker", f"{count} túra betöltve!")
            
        except Exception as e:
            QMessageBox.critical(self.dialog, "Hiba", f"Hiba a betöltés során: {e}")
        finally:
            self.main.right_tree.blockSignals(False)
            if hasattr(self.main, 'right_tree'):
                self.main.right_tree.viewport().update()

    def megjelenites(self):
        self.dialog = QDialog(self.main)
        self.dialog.setWindowTitle("Tervező Munkaablak")
        self.dialog.setMinimumSize(1200, 800)
        main_layout = QVBoxLayout(self.dialog)
        tabs = QTabWidget()
        
        tab_adatbazis = QWidget(); l1 = QVBoxLayout(tab_adatbazis)
        vezerlo_frame = QFrame(); vezerlo_frame.setStyleSheet("background-color: #f4fdf4; border: 1px solid #2ecc71; border-radius: 5px;")
        v_layout = QVBoxLayout(vezerlo_frame)

        gomb_l = QHBoxLayout()
        btn_import = QPushButton("📂 EXCEL IMPORT"); btn_import.clicked.connect(self.excel_import)
        btn_import.setStyleSheet("background-color: #34495e; color: white; font-weight: bold; padding: 10px;")
        btn_hozzaad = QPushButton("➕ RÖGZÍTÉS"); btn_hozzaad.clicked.connect(self.kezi_felvitel)
        btn_hozzaad.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 10px;")
        btn_torol = QPushButton("🗑️ TÖRLÉS"); btn_torol.clicked.connect(self.partner_torles)
        btn_torol.setStyleSheet("background-color: #c0392b; color: white; font-weight: bold; padding: 10px;")
        gomb_l.addWidget(btn_import); gomb_l.addWidget(btn_hozzaad); gomb_l.addWidget(btn_torol); gomb_l.addStretch()
        v_layout.addLayout(gomb_l)

        bevitel_l = QHBoxLayout(); self.mezok = {}
        for fejlec in self.fejlecek:
            edit = QLineEdit(); edit.setPlaceholderText(fejlec)
            if fejlec in ["Db", "Kg", "Irsz"]: edit.setFixedWidth(65)
            bevitel_l.addWidget(edit); self.mezok[fejlec] = edit
        v_layout.addLayout(bevitel_l); l1.addWidget(vezerlo_frame)

        self.tabla_p = QTableWidget(0, 8); self.tabla_p.setHorizontalHeaderLabels(self.fejlecek)
        self.tabla_p.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabla_p.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        l1.addWidget(self.tabla_p)

        tab_osszeosztas = QWidget(); l2 = QVBoxLayout(tab_osszeosztas)
        osztas_l = QHBoxLayout()
        p_vaz = QVBoxLayout(); p_vaz.addWidget(QLabel("<b>Választható Partnerek:</b>"))
        self.p_selector = QListWidget(); p_vaz.addWidget(self.p_selector); osztas_l.addLayout(p_vaz, 1)

        m_vaz = QVBoxLayout(); m_vaz.addStretch()
        btn_add = QPushButton("👉"); btn_add.setFixedSize(40, 40); btn_add.clicked.connect(self.partner_turahoz_rendel)
        btn_rem = QPushButton("🗑️"); btn_rem.setFixedSize(40, 40); btn_rem.clicked.connect(self.partner_eltavolit)
        m_vaz.addWidget(btn_add); m_vaz.addWidget(btn_rem); m_vaz.addStretch(); osztas_l.addLayout(m_vaz)

        t_vaz = QVBoxLayout(); t_vaz.addWidget(QLabel("<b>Cél Túrák:</b>"))
        self.t_selector = QComboBox()
        if os.path.exists(self.turanevek_txt):
            with open(self.turanevek_txt, "r", encoding="utf-8") as f:
                self.t_selector.addItems([line.strip() for line in f if line.strip()])
        t_vaz.addWidget(self.t_selector)
        self.t_tartalom = QListWidget(); t_vaz.addWidget(self.t_tartalom)
        self.t_selector.currentTextChanged.connect(self.tura_tartalom_mutat); osztas_l.addLayout(t_vaz, 1)
        l2.addLayout(osztas_l)

        # JAVÍTOTT FELIRAT
        self.btn_betolt = QPushButton("🚀 ÖSSZEOSZTOTT TÚRÁK BETÖLTÉSE A FŐPROGRAMBA")
        self.btn_betolt.setStyleSheet("background-color: #2980b9; color: white; font-weight: bold; padding: 15px;")
        self.btn_betolt.clicked.connect(self.turak_behelyezese_a_foprogramba)
        l2.addWidget(self.btn_betolt)

        tabs.addTab(tab_adatbazis, "1. Adatbázis"); tabs.addTab(tab_osszeosztas, "2. Összeosztás")
        main_layout.addWidget(tabs); self.frissit_felulet(); self.dialog.exec()

    def kezi_felvitel(self):
        uj = {f: self.mezok[f].text().strip() for f in self.fejlecek}
        if not uj["Név"]: return
        self.partnerek.append({"nev": uj["Név"], "ki": uj["KI"], "be": uj["BE"], "db": uj["Db"] or "0", 
                               "kg": uj["Kg"] or "0", "megj": uj["Megjegyzés"], "varos": uj["Város"], "irsz": uj["Irsz"]})
        self.mentes_adatok(); self.frissit_felulet()
        for m in self.mezok.values(): m.clear()

    def partner_torles(self):
        row = self.tabla_p.currentRow()
        if row >= 0:
            self.partnerek.pop(row); self.mentes_adatok(); self.frissit_felulet()

    def excel_import(self):
        fajl, _ = QFileDialog.getOpenFileName(self.dialog, "Excel", "", "Excel fájlok (*.xlsx *.xls)")
        if not fajl: return
        try:
            wb = load_workbook(fajl, data_only=True); ws = wb.active; volt_uj = 0
            for r in range(2, ws.max_row + 1):
                nev = str(ws.cell(row=r, column=3).value or "").strip()
                if nev and nev != "None":
                    uj = {"nev": nev, "ki": str(ws.cell(row=r, column=5).value or ""), 
                          "be": str(ws.cell(row=r, column=6).value or ""), "db": str(ws.cell(row=r, column=7).value or "0"),
                          "kg": str(ws.cell(row=r, column=8).value or "0"), "megj": str(ws.cell(row=r, column=4).value or ""),
                          "varos": str(ws.cell(row=r, column=10).value or ""), "irsz": str(ws.cell(row=r, column=9).value or "")}
                    if not any(p['nev'] == uj['nev'] for p in self.partnerek):
                        self.partnerek.append(uj); volt_uj += 1
            if volt_uj > 0: self.mentes_adatok(); self.frissit_felulet()
            QMessageBox.information(self.dialog, "Kész", f"Importálva: {volt_uj}")
        except Exception as e: QMessageBox.critical(self.dialog, "Hiba", str(e))

    def frissit_felulet(self):
        self.tabla_p.setRowCount(0); self.p_selector.clear()
        for i, p in enumerate(self.partnerek):
            self.tabla_p.insertRow(i)
            sor = [p.get("nev",""), p.get("ki",""), p.get("be",""), p.get("db","0"), p.get("kg","0"), p.get("megj",""), p.get("varos",""), p.get("irsz","")]
            for c, text in enumerate(sor): self.tabla_p.setItem(i, c, QTableWidgetItem(str(text)))
            self.p_selector.addItem(p["nev"])
        self.tura_tartalom_mutat(self.t_selector.currentText())

    def tura_tartalom_mutat(self, t_nev):
        self.t_tartalom.clear()
        if t_nev in self.tura_kapcsolatok: self.t_tartalom.addItems(self.tura_kapcsolatok[t_nev])

    def partner_turahoz_rendel(self):
        p = self.p_selector.currentItem(); t = self.t_selector.currentText()
        if p and t:
            if t not in self.tura_kapcsolatok: self.tura_kapcsolatok[t] = []
            if p.text() not in self.tura_kapcsolatok[t]:
                self.tura_kapcsolatok[t].append(p.text()); self.mentes_adatok(); self.tura_tartalom_mutat(t)

    def partner_eltavolit(self):
        p = self.t_tartalom.currentItem(); t = self.t_selector.currentText()
        if p and t in self.tura_kapcsolatok:
            self.tura_kapcsolatok[t].remove(p.text()); self.mentes_adatok(); self.tura_tartalom_mutat(t)
